# app_dashboard.py
import os
import sys
import platform
import subprocess
import webbrowser
from pathlib import Path

def setup_environment():
    """Configura o ambiente automaticamente para Windows ou Linux"""
    system = platform.system()
    
    print(f"🔍 Detectado: {system}")
    
    # Define o diretório base
    if system == "Windows":
        base_dir = Path(os.getcwd())
    else:  # Linux
        base_dir = Path(os.getcwd())
    
    # Lista de dependências
    requirements = ["pandas", "openpyxl", "xlrd", "customtkinter", "pillow"]
    
    # Tenta importar ou instala as dependências
    for package in requirements:
        try:
            __import__(package)
            print(f"✅ {package} já instalado")
        except ImportError:
            print(f"📦 Instalando {package}...")
            try:
                subprocess.check_call([sys.executable, "-m", "pip", "install", package])
                print(f"✅ {package} instalado com sucesso")
            except subprocess.CalledProcessError:
                print(f"❌ Falha ao instalar {package}")
    
    return base_dir

def main():
    """Função principal que inicia a aplicação"""
    try:
        base_dir = setup_environment()
        
        # Agora importa as bibliotecas
        import pandas as pd
        from openpyxl import load_workbook
        import customtkinter as ctk
        from tkinter import messagebox, filedialog
        import threading
        import glob
        import shutil
        from datetime import datetime
        import csv
        
    except Exception as e:
        print(f"❌ Erro crítico: {e}")
        input("Pressione Enter para sair...")
        return

    # Mapeamentos (mantidos)
    CATEGORY_MAPPING = {
        'DESPESAS ADMINISTRATIVAS : Água e Esgoto': 'Água e Esgoto',
        'DESPESAS ADMINISTRATIVAS : Energia Elétrica': 'Energia Elétrica',
        'DESPESAS ADMINISTRATIVAS : Internet': 'Internet',
        'DESPESAS ADMINISTRATIVAS : Sistemas e Softwares': 'Sistemas e Softwares',
        'DESPESAS ADMINISTRATIVAS : Telefonia': 'Telefonia',
        'DESPESA COM PESSOAL : Benefícios - Vale-Transporte': 'Beneficios - Vale-Transporte',
        'DESPESA COM PESSOAL : Benefícios Alimentação e Refeição': 'Beneficios - Vale-Alimentação',
        'DESPESA COM PESSOAL : Encargos Sociais - FGTS': 'Encargos Sociais - FGTS',
        'DESPESA COM PESSOAL : Encargos Sociais - INSS': 'Encargos Sociais - INSS',
        'DESPESA COM PESSOAL : Estagiários (bolsa + seguro)': 'Estagiários (bolsa + seguro)',
        'DESPESA COM PESSOAL : Rescisões': 'Rescisões',
        'DESPESA COM PESSOAL : Salários': 'Salários',
        'DESPESA COM PESSOAL : SST (Segurança e Saúde do Trabalho)': 'SST (Segurança e Saúde do Trabalho)',
        'DESPESAS FINANCEIRAS : Tarifas Bancárias - Boletos': 'Tarifas Bancárias - Boletos',
        'DESPESAS FINANCEIRAS : Tarifas Bancárias - Cartão': 'Tarifas Cartão de Crédito',
        'DESPESAS FINANCEIRAS : Tarifas Bancárias - Pix': 'Tarifas Bancárias – Pix',
        'DESPESAS TRIBUTÁRIAS : SIMPLES NACIONAL': 'SIMPLES',
        'DESPESAS TRIBUTÁRIAS : Taxas Municipais (Alvarás, Bombeiros e etc.)': 'Taxas Municipais (Alvarás, Bombeiros etc.)',
        'ADIANTAMENTO DIVIDENDOS : Adiantamento de Sócio': 'Adiantamento de Sócio',
        'SERVIÇOS PROFISSIONAIS E LEGAIS : Honorários Contábeis': 'Honorários Contábeis',
        'INVESTIMENTO E MANUTENÇÕES : Aquisição de Equipamentos': 'Aquisição de Equipamentos',
        'INVESTIMENTO E MANUTENÇÕES : Aquisição de Utensílios e Componentes': 'Aquisição de Utensílios e Componentes'
    }

    SAO_PAULO_MAPPING = {
        'ADIANTAMENTO DIVIDENDOS:Adiantamento de Sócio': 'Adiantamento de Sócio',
        'DESPESAS ADMINISTRATIVAS:Internet': 'Internet',
        'DESPESAS ADMINISTRATIVAS:Sistemas e Softwares': 'Sistemas e Softwares',
        'DESPESAS ADMINISTRATIVAS:Telefonia': 'Telefonia',
        'DESPESAS FINANCEIRAS:Tarifas Bancárias - Boletos': 'Tarifas Bancárias - Boletos',
        'DESPESAS FINANCEIRAS:Tarifas Bancárias - Cartão': 'Tarifas Cartão de Crédito',
        'DESPESAS FINANCEIRAS:Tarifas Bancárias - Pix': 'Tarifas Bancárias – Pix',
        'DESPESAS TRIBUTÁRIAS:SIMPLES NACIONAL': 'SIMPLES',
        'SERVIÇOS PROFISSIONAIS E LEGAIS:Honorários Contábeis': 'Honorários Contábeis',
        'INVESTIMENTO E MANUTENÇÕES:Aquisição de Equipamentos': 'Aquisição de Equipamentos',
        'INVESTIMENTO E MANUTENÇÕES:Aquisição de Utensílios e Componentes': 'Aquisição de Utensílios e Componentes'
    }

    class DashboardApp(ctk.CTk):
        def __init__(self):
            super().__init__()
            
            # Configuração da janela
            self.title("Athena Office - Transporte de Dados")
            self.geometry("1000x700")
            self.minsize(900, 600)
            
            # Centraliza a janela
            self.center_window()
            
            # Variáveis
            self.base_dir = base_dir
            self.final_name = "DASHBOARDFINAL.xlsx"
            self.final_path = self.base_dir / self.final_name
            self.is_processing = False
            
            self.setup_ui()
            self.update_initial_info()
            
        def center_window(self):
            """Centraliza a janela na tela"""
            self.update_idletasks()
            width = self.winfo_width()
            height = self.winfo_height()
            x = (self.winfo_screenwidth() // 2) - (width // 2)
            y = (self.winfo_screenheight() // 2) - (height // 2)
            self.geometry(f"{width}x{height}+{x}+{y}")
            
        def setup_ui(self):
            # Configuração do tema - Windows-friendly
            ctk.set_appearance_mode("Light")
            ctk.set_default_color_theme("blue")
            
            # Layout principal
            self.grid_columnconfigure(1, weight=1)
            self.grid_rowconfigure(0, weight=1)
            
            # Sidebar
            self.sidebar_frame = ctk.CTkFrame(self, width=200, corner_radius=0)
            self.sidebar_frame.grid(row=0, column=0, sticky="nsew")
            self.sidebar_frame.grid_rowconfigure(6, weight=1)
            
            # Logo
            self.logo_label = ctk.CTkLabel(
                self.sidebar_frame, 
                text="Athena Office",
                font=ctk.CTkFont(size=20, weight="bold")
            )
            self.logo_label.grid(row=0, column=0, padx=20, pady=20)
            
            self.subtitle_label = ctk.CTkLabel(
                self.sidebar_frame,
                text="Transporte de Dados",
                font=ctk.CTkFont(size=12)
            )
            self.subtitle_label.grid(row=1, column=0, padx=20, pady=(0, 20))
            
            # Botões principais
            self.process_btn = ctk.CTkButton(
                self.sidebar_frame,
                text="🚀 Processar Dados",
                command=self.start_processing,
                font=ctk.CTkFont(size=14, weight="bold"),
                height=40
            )
            self.process_btn.grid(row=2, column=0, padx=20, pady=10)
            
            self.select_folder_btn = ctk.CTkButton(
                self.sidebar_frame,
                text="📁 Selecionar Pasta",
                command=self.select_folder,
                font=ctk.CTkFont(size=12)
            )
            self.select_folder_btn.grid(row=3, column=0, padx=20, pady=5)
            
            self.help_btn = ctk.CTkButton(
                self.sidebar_frame,
                text="❓ Ajuda",
                command=self.show_help,
                font=ctk.CTkFont(size=12)
            )
            self.help_btn.grid(row=4, column=0, padx=20, pady=5)
            
            # Info do sistema
            self.system_label = ctk.CTkLabel(
                self.sidebar_frame,
                text=f"Sistema: {platform.system()}",
                font=ctk.CTkFont(size=10)
            )
            self.system_label.grid(row=7, column=0, padx=20, pady=10)
            
            # Área principal
            self.main_frame = ctk.CTkFrame(self, corner_radius=10)
            self.main_frame.grid(row=0, column=1, sticky="nsew", padx=20, pady=20)
            self.main_frame.grid_columnconfigure(0, weight=1)
            self.main_frame.grid_rowconfigure(1, weight=1)
            
            # Título
            self.main_title = ctk.CTkLabel(
                self.main_frame,
                text="Sistema de Transporte de Dados",
                font=ctk.CTkFont(size=20, weight="bold")
            )
            self.main_title.grid(row=0, column=0, padx=20, pady=20)
            
            # Área de status
            self.status_frame = ctk.CTkFrame(self.main_frame)
            self.status_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=(0, 20))
            self.status_frame.grid_columnconfigure(0, weight=1)
            self.status_frame.grid_rowconfigure(1, weight=1)
            
            # Texto de informações
            self.info_text = ctk.CTkTextbox(self.status_frame, height=200)
            self.info_text.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
            self.info_text.configure(state="disabled")
            
            # Progresso
            self.progress_frame = ctk.CTkFrame(self.status_frame)
            self.progress_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=10)
            self.progress_frame.grid_columnconfigure(0, weight=1)
            
            self.progress_label = ctk.CTkLabel(
                self.progress_frame,
                text="Pronto para processar",
                font=ctk.CTkFont(size=12)
            )
            self.progress_label.grid(row=0, column=0, sticky="w", padx=10, pady=5)
            
            self.progress_bar = ctk.CTkProgressBar(self.progress_frame)
            self.progress_bar.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 10))
            self.progress_bar.set(0)
            
            # Estatísticas
            self.stats_frame = ctk.CTkFrame(self.main_frame)
            self.stats_frame.grid(row=2, column=0, sticky="ew", padx=20, pady=(0, 20))
            self.stats_frame.grid_columnconfigure((0, 1, 2), weight=1)
            
            self.cities_label = ctk.CTkLabel(
                self.stats_frame,
                text="Cidades: 0",
                font=ctk.CTkFont(size=14, weight="bold")
            )
            self.cities_label.grid(row=0, column=0, padx=20, pady=15)
            
            self.updated_label = ctk.CTkLabel(
                self.stats_frame,
                text="Atualizadas: 0",
                font=ctk.CTkFont(size=14, weight="bold")
            )
            self.updated_label.grid(row=0, column=1, padx=20, pady=15)
            
            self.categories_label = ctk.CTkLabel(
                self.stats_frame,
                text="Categorias: 0",
                font=ctk.CTkFont(size=14, weight="bold")
            )
            self.categories_label.grid(row=0, column=2, padx=20, pady=15)
        
        def select_folder(self):
            """Permite ao usuário selecionar uma pasta diferente"""
            folder = filedialog.askdirectory(title="Selecione a pasta com os arquivos")
            if folder:
                self.base_dir = Path(folder)
                self.final_path = self.base_dir / self.final_name
                self.update_initial_info()
                messagebox.showinfo("Sucesso", f"Pasta selecionada:\n{self.base_dir}")
        
        def show_help(self):
            """Mostra janela de ajuda"""
            help_text = """
🎯 MANUAL RÁPIDO ATHENA DASHBOARD

📋 COMO USAR:
1. Coloque todos os arquivos na mesma pasta:
   - DASHBOARDFINAL.xlsx (Dashboard principal)
   - SãoPaulo.xls, JoãoPessoa.xls, etc. (Planilhas das cidades)

2. Clique em "Processar Dados"

3. Aguarde o processamento

🛡️ RECURSOS:
- Backup automático antes de alterações
- Detecta automaticamente os arquivos
- Interface simples e intuitiva

📁 ESTRUTURA:
Pasta/
├── DASHBOARDFINAL.xlsx
├── SãoPaulo.xls
├── JoãoPessoa.xls
└── app_dashboard.py

❓ PROBLEMAS COMUNS:
- Verifique se o Excel não está aberto
- Todos arquivos devem estar na mesma pasta
- Nomes dos arquivos devem corresponder às abas

💡 DICA: Use o botão "Selecionar Pasta" se os arquivos estiverem em outra localização
            """
            messagebox.showinfo("Ajuda - Athena Dashboard", help_text)
        
        def update_initial_info(self):
            self.info_text.configure(state="normal")
            self.info_text.delete("1.0", "end")
            
            info_lines = [
                "📊 SISTEMA DE TRANSPORTE DE DADOS",
                "=" * 50,
                f"📁 Pasta: {self.base_dir}",
                f"💻 Sistema: {platform.system()}",
                "",
                "🔍 ARQUIVOS ENCONTRADOS:"
            ]
            
            # Verifica se o dashboard existe
            if not self.final_path.exists():
                info_lines.append("❌ DASHBOARDFINAL.xlsx NÃO ENCONTRADO!")
            else:
                info_lines.append("✅ DASHBOARDFINAL.xlsx - OK")
            
            city_files = self.find_city_files()
            for city_file in city_files:
                info_lines.append(f"   📄 {os.path.basename(city_file)}")
            
            if not city_files:
                info_lines.append("   ⚠️  Nenhuma planilha de cidade encontrada")
            
            info_lines.extend([
                "",
                "🎯 INSTRUÇÕES:",
                "1. Clique em 'Processar Dados' para iniciar",
                "2. Aguarde o processamento automático", 
                "3. Backup será criado automaticamente",
                "",
                "✅ PRONTO PARA USAR" if self.final_path.exists() else "❌ CONFIGURE OS ARQUIVOS PRIMEIRO"
            ])
            
            self.info_text.insert("1.0", "\n".join(info_lines))
            self.info_text.configure(state="disabled")
            self.cities_label.configure(text=f"Cidades: {len(city_files)}")
        
        def find_city_files(self):
            patterns = ["*.xls", "*.xlsx"]
            files = []
            for pattern in patterns:
                files.extend(glob.glob(str(self.base_dir / pattern)))
            
            return [
                f for f in files
                if os.path.basename(f) != self.final_name
                and not os.path.basename(f).startswith("~$")
                and not os.path.basename(f).startswith("DASHBOARDFINAL_backup_")
            ]
        
        def start_processing(self):
            if self.is_processing:
                return
            
            # Verifica se o dashboard existe
            if not self.final_path.exists():
                messagebox.showerror("Erro", f"Arquivo não encontrado:\n{self.final_path}\n\nColoque o arquivo DASHBOARDFINAL.xlsx na pasta.")
                return
                
            self.is_processing = True
            self.process_btn.configure(state="disabled")
            self.progress_bar.set(0)
            
            self.info_text.configure(state="normal")
            self.info_text.delete("1.0", "end")
            self.info_text.insert("1.0", "🚀 INICIANDO PROCESSAMENTO...\n\n")
            self.info_text.configure(state="disabled")
            
            thread = threading.Thread(target=self.process_data)
            thread.daemon = True
            thread.start()
        
        def process_data(self):
            try:
                self.log_message("📁 Criando backup do dashboard...")
                
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_path = self.base_dir / f"DASHBOARDFINAL_backup_{ts}.xlsx"
                
                if self.final_path.exists():
                    shutil.copyfile(self.final_path, backup_path)
                    self.log_message(f"✅ Backup criado: {backup_path.name}")
                else:
                    self.log_message("❌ Dashboard não encontrado!")
                    self.finish_processing(False)
                    return
                
                self.log_message("📊 Carregando dashboard...")
                try:
                    wb = load_workbook(self.final_path)
                    self.log_message(f"✅ Dashboard carregado - {len(wb.sheetnames)} abas")
                except Exception as e:
                    self.log_message(f"❌ Erro ao carregar dashboard: {e}")
                    self.finish_processing(False)
                    return
                
                city_files = self.find_city_files()
                self.log_message(f"🏙️  {len(city_files)} arquivos de cidades encontrados")
                
                if not city_files:
                    self.log_message("⚠️  Nenhuma planilha de cidade para processar")
                    self.finish_processing(True)
                    return
                
                total_updated = 0
                total_categories = 0
                
                for i, city_file in enumerate(city_files):
                    progress = (i / len(city_files)) * 100
                    self.update_progress(progress, f"Processando {os.path.basename(city_file)}...")
                    
                    city_name = os.path.splitext(os.path.basename(city_file))[0]
                    self.log_message(f"\n📋 PROCESSANDO: {city_name}")
                    
                    try:
                        if "são paulo" in city_name.lower() or "sao paulo" in city_name.lower():
                            expenses = self.extract_expenses_sao_paulo(city_file)
                        else:
                            expenses = self.extract_expenses_joao_pessoa(city_file)
                        
                        if expenses:
                            updated = self.update_dashboard_city_sheet(wb, city_name, expenses)
                            total_updated += updated
                            total_categories += len(expenses)
                            
                            self.log_message(f"✅ {city_name}: {len(expenses)} categorias, {updated} atualizadas")
                        else:
                            self.log_message(f"⚠️  {city_name}: Nenhum dado extraído")
                            
                    except Exception as e:
                        self.log_message(f"❌ Erro em {city_name}: {e}")
                
                self.update_progress(90, "Salvando dashboard...")
                try:
                    wb.save(self.final_path)
                    self.log_message(f"💾 Dashboard salvo com sucesso!")
                except Exception as e:
                    self.log_message(f"❌ Erro ao salvar: {e}")
                    self.finish_processing(False)
                    return
                
                self.after(0, lambda: self.update_final_stats(total_updated, total_categories, len(city_files)))
                
                self.update_progress(100, "Processamento concluído!")
                self.log_message(f"\n🎯 PROCESSAMENTO CONCLUÍDO!")
                self.log_message(f"📊 Resumo: {len(city_files)} cidades, {total_categories} categorias, {total_updated} atualizações")
                
                self.finish_processing(True)
                
            except Exception as e:
                self.log_message(f"❌ ERRO CRÍTICO: {e}")
                self.finish_processing(False)
        
        def log_message(self, message):
            def update_text():
                self.info_text.configure(state="normal")
                self.info_text.insert("end", message + "\n")
                self.info_text.see("end")
                self.info_text.configure(state="disabled")
            
            self.after(0, update_text)
        
        def update_progress(self, value, text):
            def update():
                self.progress_bar.set(value / 100)
                self.progress_label.configure(text=text)
            
            self.after(0, update)
        
        def update_final_stats(self, updated, categories, cities):
            self.updated_label.configure(text=f"Atualizadas: {updated}")
            self.categories_label.configure(text=f"Categorias: {categories}")
            self.cities_label.configure(text=f"Cidades: {cities}")
        
        def finish_processing(self, success):
            def finish():
                self.is_processing = False
                self.process_btn.configure(state="normal")
                
                if success:
                    messagebox.showinfo("Sucesso", "Processamento concluído com sucesso!\n\nUm backup foi criado automaticamente.")
                else:
                    messagebox.showerror("Erro", "Ocorreu um erro durante o processamento.\n\nVerifique se o Excel não está aberto.")
            
            self.after(0, finish)
        
        # Funções de processamento (mantidas)
        def extract_expenses_joao_pessoa(self, file_path):
            try:
                df = pd.read_excel(file_path, engine='xlrd')
                expenses = {}
                
                for _, row in df.iterrows():
                    if len(row) >= 2 and pd.notna(row[0]) and pd.notna(row[1]):
                        descricao = str(row[0]).strip()
                        valor = row[1]
                        
                        try:
                            valor_numerico = float(valor)
                            
                            for cat_origem, cat_destino in CATEGORY_MAPPING.items():
                                if cat_origem in descricao:
                                    expenses[cat_destino] = valor_numerico
                                    self.log_message(f"  ✅ {cat_destino}: R$ {valor_numerico:,.2f}")
                                    break
                            
                        except (ValueError, TypeError):
                            continue
                
                return expenses
            
            except Exception as e:
                self.log_message(f"❌ Erro ao processar: {e}")
                return {}

        def extract_expenses_sao_paulo(self, file_path):
            try:
                df = pd.read_excel(file_path, engine='xlrd', header=None)
                expenses = {}
                
                if df.shape[0] >= 2:
                    for col in range(1, df.shape[1]):
                        categoria = str(df.iloc[0, col]).strip() if pd.notna(df.iloc[0, col]) else ''
                        valor = df.iloc[1, col] if pd.notna(df.iloc[1, col]) else None
                        
                        if categoria and valor is not None:
                            try:
                                valor_numerico = float(valor)
                                
                                for cat_origem, cat_destino in SAO_PAULO_MAPPING.items():
                                    if cat_origem in categoria:
                                        expenses[cat_destino] = valor_numerico
                                        self.log_message(f"  ✅ {cat_destino}: R$ {valor_numerico:,.2f}")
                                        break
                                
                            except (ValueError, TypeError):
                                continue
                
                return expenses
            
            except Exception as e:
                self.log_message(f"❌ Erro ao processar: {e}")
                return {}

        def find_city_sheet(self, wb, city_name):
            city_lower = city_name.lower()
            
            for sheet_name in wb.sheetnames:
                if city_lower in sheet_name.lower():
                    return sheet_name
            
            return None

        def update_dashboard_city_sheet(self, wb, city_name, expenses):
            sheet_name = self.find_city_sheet(wb, city_name)
            
            if not sheet_name:
                self.log_message(f"❌ Aba para '{city_name}' não encontrada")
                return 0
            
            ws = wb[sheet_name]
            updated_count = 0
            
            for row in range(1, ws.max_row + 1):
                cell_a = ws[f'A{row}']
                
                if cell_a.value and isinstance(cell_a.value, str):
                    categoria_dashboard = cell_a.value.strip()
                    
                    for cat_extraida, valor in expenses.items():
                        if cat_extraida.lower() in categoria_dashboard.lower():
                            cell_b = ws[f'B{row}']
                            old_value = cell_b.value
                            cell_b.value = valor
                            updated_count += 1
                            
                            if old_value != valor:
                                self.log_message(f"  🔄 {categoria_dashboard}: R$ {valor:,.2f} (era: {old_value})")
                            break
            
            return updated_count

    # Inicia a aplicação
    app = DashboardApp()
    app.mainloop()

if __name__ == "__main__":
    main()