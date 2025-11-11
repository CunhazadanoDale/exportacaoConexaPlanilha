# -*- coding: utf-8 -*-
import os
import sys
import re
import unicodedata
import platform
import subprocess
from pathlib import Path
from difflib import SequenceMatcher

def setup_environment():
    system = platform.system()
    print(f"🔍 Detectado: {system}")
    base_dir = Path.cwd()

    requirements = [
        ("pandas", "pandas"),
        ("openpyxl", "openpyxl"),
        ("xlrd", "xlrd"),
        ("customtkinter", "customtkinter"),
        ("pillow", "PIL"),
    ]
    is_frozen = getattr(sys, "frozen", False)
    missing = []
    for pip_name, import_name in requirements:
        try:
            __import__(import_name)
            print(f"✅ {import_name} disponível")
        except ImportError:
            if is_frozen:
                missing.append((pip_name, import_name))
            else:
                print(f"📦 Instalando {pip_name}...")
                try:
                    subprocess.check_call([sys.executable, "-m", "pip", "install", pip_name])
                    print(f"✅ {pip_name} instalado com sucesso")
                except subprocess.CalledProcessError:
                    print(f"❌ Falha ao instalar {pip_name}")
                    missing.append((pip_name, import_name))
    if missing:
        msg = "Dependências ausentes: " + ", ".join([f"{imp}(pip:{pip})" for pip, imp in missing])
        if is_frozen:
            msg += ("\n\nEste executável não instala pacotes em runtime.\n"
                    "Reempacote incluindo as dependências (veja requirements.txt) "
                    "ou rode via Python com `pip install -r requirements.txt`.")
        raise RuntimeError(msg)
    return base_dir

# --------- Normalização e helpers ---------
def strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

def norm_text(s: str) -> str:
    # lower, remove acentos, tira espaços extras e espaços ao redor de ':'
    s = str(s)
    # Substitui caracteres especiais antes de remover acentos
    s = s.replace('–', '-').replace('—', '-').replace('―', '-')
    s = s.replace('´', "'").replace('`', "'").replace('’', "'").replace('‘', "'")
    s = strip_accents(s).lower()
    s = re.sub(r"\s*:\s*", ":", s)          # " : " -> ":"
    s = re.sub(r"\s+", " ", s).strip()      # colapsa espaços
    return s

def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            return float(v)
        except Exception:
            return None
    s = str(v).strip()
    if not s:
        return None
    s = re.sub(r"[Rr]\$?\s*", "", s)
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    s = s.replace(" ", "")
    try:
        return float(s)
    except Exception:
        return None

def main():
    try:
        base_dir = setup_environment()
        import pandas as pd
        from openpyxl import load_workbook
        import customtkinter as ctk
        from tkinter import messagebox, filedialog
        import threading, glob, shutil
        from datetime import datetime
    except Exception as e:
        print(f"❌ Erro crítico: {e}")
        try:
            input("Pressione Enter para sair...")
        except Exception:
            pass
        return

    # MAPEAMENTO COMPLETO E CORRIGIDO
    RAW_CATEGORY_MAPPING = {
        # DESPESAS ADMINISTRATIVAS
        'DESPESAS ADMINISTRATIVAS : Água e Esgoto': 'Água e Esgoto',
        'DESPESAS ADMINISTRATIVAS : Aluguel e Condomínio': 'Aluguel e Condominio',
        'DESPESAS ADMINISTRATIVAS : Energia Elétrica': 'Energia Elétrica',
        'DESPESAS ADMINISTRATIVAS : Internet': 'Internet',
        'DESPESAS ADMINISTRATIVAS : Sistemas e Softwares': 'Sistemas e Softwares',
        'DESPESAS ADMINISTRATIVAS : Telefonia': 'Telefonia',
        'DESPESAS ADMINISTRATIVAS : Transporte / Deslocamentos (Uber, táxi, viagens administrativas)': 'Transporte / Deslocamentos (Uber, táxi, viagens administrativas)',
        'DESPESAS ADMINISTRATIVAS : Serviços Terceirizados (Limpeza, segurança, etc.)': 'Serviços Terceirizados (Limpeza, segurança, etc.)',
        'DESPESAS ADMINISTRATIVAS : Materiais de Escritório': 'Materiais de Escritório',
        'DESPESAS ADMINISTRATIVAS : Materiais de Limpeza': 'Materiais de Limpeza',
        'DESPESAS ADMINISTRATIVAS : Materiais de Copa/Cozinha': 'Materiais de Copa/Cozinha',
        'DESPESAS ADMINISTRATIVAS : Consultoria Externa': 'Consultoria Externa',
        'DESPESAS ADMINISTRATIVAS : Serviços Postais / Correspondência': 'Serviços Postais / Correspondência',
        'DESPESAS ADMINISTRATIVAS : Serviços Cartorários': 'Serviços Cartorários',
        
        # DESPESA COM PESSOAL
        'DESPESA COM PESSOAL : Benefícios - Vale-Transporte': 'Beneficios - Vale-Transporte',
        'DESPESA COM PESSOAL : Benefícios Alimentação e Refeição': 'Beneficios - Vale-Alimentação',
        'DESPESA COM PESSOAL : Encargos Sociais - FGTS': 'Encargos Sociais - FGTS',
        'DESPESA COM PESSOAL : Encargos Sociais - INSS': 'Encargos Sociais - INSS',
        'DESPESA COM PESSOAL : Estagiários (bolsa + seguro)': 'Estagiários (bolsa + seguro)',
        'DESPESA COM PESSOAL : Rescisões': 'Rescisões',
        'DESPESA COM PESSOAL : Salários': 'Salários',
        'DESPESA COM PESSOAL : SST (Segurança e Saúde do Trabalho)': 'SST (Segurança e Saúde do Trabalho)',
        'DESPESA COM PESSOAL : Multas Rescisórias': 'Multas Rescisórias',
        'DESPESA COM PESSOAL : Cursos e Treinamentos': 'Cursos e Treinamentos',
        'DESPESA COM PESSOAL : Fardamentos': 'Fardamentos',
        'DESPESA COM PESSOAL : Outros custos com Pessoal': 'Outros custos com Pessoal',
        
        # DESPESAS FINANCEIRAS
        'DESPESAS FINANCEIRAS : Juros e Multas': 'Juros e Multas',
        'DESPESAS FINANCEIRAS : Tarifas Bancárias - Boletos': 'Tarifas Bancárias - Boletos',
        'DESPESAS FINANCEIRAS : Tarifas Bancárias - Cartão': 'Tarifas Cartão Crédito',
        'DESPESAS FINANCEIRAS : Tarifas Bancárias - Pix': 'Tarifas Bancárias – Pix',
        'DESPESAS FINANCEIRAS : Tarifas Bancárias - TED': 'Tarifas Bancárias TED',
        'DESPESAS FINANCEIRAS : Empréstimos e Financiamentos': 'Empréstimos e Financiamentos',
        
        # DESPESAS TRIBUTÁRIAS
        'DESPESAS TRIBUTÁRIAS : SIMPLES NACIONAL': 'SIMPLES',
        'DESPESAS TRIBUTÁRIAS : Taxas Municipais (Alvarás, Bombeiros e etc.)': 'Taxas Municipais (Alvarás, Bombeiros etc.)',
        'DESPESAS TRIBUTÁRIAS : IPTU': 'IPTU',
        'DESPESAS TRIBUTÁRIAS : ISS': 'ISS',
        'DESPESAS TRIBUTÁRIAS : Outras Taxas e Contribuições': 'Outras Taxas e Contribuições',
        
        # OUTRAS CATEGORIAS
        'ADIANTAMENTO DIVIDENDOS : Adiantamento de Sócio': 'Adiantamento de Sócio',
        'SERVIÇOS PROFISSIONAIS E LEGAIS : Honorários Contábeis': 'Honorários Contábeis',
        'SERVIÇOS PROFISSIONAIS E LEGAIS : Honorários JuridicoS': 'Honorários JuridicoS',
        'SERVIÇOS PROFISSIONAIS E LEGAIS : Despesas Legais e Judiciais': 'Despesas Legais e Judiciais',
        'INVESTIMENTO E MANUTENÇÕES : Aquisição de Equipamentos': 'Aquisição de Equipamentos',
        'INVESTIMENTO E MANUTENÇÕES : Aquisição de Móveis': 'Aquisição de Móveis',
        'INVESTIMENTO E MANUTENÇÕES : Aquisição de Utensílios e Componentes': 'Aquisição de Utensílios e Componentes',
        'INVESTIMENTO E MANUTENÇÕES : Manutenção de Equipamentos': 'Manutenção de Equipamentos',
        'INVESTIMENTO E MANUTENÇÕES : Manutenção do Escritório (mobiliário, infraestrutura, elétrica etc.)': 'Manutenção do Escritório (mobiliário, infraestrutura, elétrica etc.)',
        'MARKETING E COMUNICAÇÃO : Propaganda e Publicidade': 'Propaganda e Publicidade',

        'AJUSTES E REGULARIZAÇÕES : Devoluções': 'Devoluções'
    }
    
    # versão normalizada do mapping
    CATEGORY_MAPPING = {norm_text(k): v for k, v in RAW_CATEGORY_MAPPING.items()}

    # --------- leitor resiliente ---------
    def read_excel_any(path: Path):
        import pandas as pd

        # Fareja assinatura do arquivo
        with open(path, "rb") as f:
            sig = f.read(8)

        is_zip = sig.startswith(b"PK")  # .xlsx (zip)
        is_ole = sig.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1")  # .xls (OLE)

        # 1) Arquivo é ZIP -> tratar como XLSX com openpyxl (mesmo que a extensão seja .xls)
        if is_zip:
            return pd.read_excel(path, engine="openpyxl")

        # 2) Arquivo é OLE -> tratar como XLS com xlrd
        if is_ole:
            return pd.read_excel(path, engine="xlrd")

        # 3) Desconhecido: tenta auto, depois openpyxl, depois xlrd; se nada, orienta a salvar como xlsx
        try:
            return pd.read_excel(path)
        except Exception:
            pass
        try:
            return pd.read_excel(path, engine="openpyxl")
        except Exception:
            pass
        try:
            return pd.read_excel(path, engine="xlrd")
        except Exception as e:
            raise RuntimeError(
                f"Falha ao ler '{Path(path).name}': {e}\n"
                "O arquivo parece estar com formato/assinatura inconsistente. "
                "Abra e 'Salvar como' .xlsx, depois rode novamente."
            )


    class DashboardApp(ctk.CTk):
        def __init__(self):
            super().__init__()
            self.title("Athena Office - Transporte de Dados")
            self.geometry("1000x700")
            self.minsize(900, 600)
            self.center_window()

            self.base_dir = base_dir
            self.final_name = "DASHBOARDFINAL.xlsx"
            self.final_path = self.base_dir / self.final_name
            self.is_processing = False

            self.setup_ui()
            self.update_initial_info()

        def center_window(self):
            self.update_idletasks()
            width = self.winfo_width(); height = self.winfo_height()
            x = (self.winfo_screenwidth() // 2) - (width // 2)
            y = (self.winfo_screenheight() // 2) - (height // 2)
            self.geometry(f"{width}x{height}+{x}+{y}")

        def setup_ui(self):
            import customtkinter as ctk
            ctk.set_appearance_mode("Light")
            ctk.set_default_color_theme("blue")

            self.grid_columnconfigure(1, weight=1)
            self.grid_rowconfigure(0, weight=1)

            self.sidebar_frame = ctk.CTkFrame(self, width=200, corner_radius=0)
            self.sidebar_frame.grid(row=0, column=0, sticky="nsew")
            self.sidebar_frame.grid_rowconfigure(6, weight=1)

            self.logo_label = ctk.CTkLabel(self.sidebar_frame, text="Athena Office", font=ctk.CTkFont(size=20, weight="bold"))
            self.logo_label.grid(row=0, column=0, padx=20, pady=20)

            self.subtitle_label = ctk.CTkLabel(self.sidebar_frame, text="Transporte de Dados", font=ctk.CTkFont(size=12))
            self.subtitle_label.grid(row=1, column=0, padx=20, pady=(0, 20))

            self.process_btn = ctk.CTkButton(self.sidebar_frame, text="🚀 Processar Dados", command=self.start_processing, font=ctk.CTkFont(size=14, weight="bold"), height=40)
            self.process_btn.grid(row=2, column=0, padx=20, pady=10)

            self.select_folder_btn = ctk.CTkButton(self.sidebar_frame, text="📁 Selecionar Pasta", command=self.select_folder, font=ctk.CTkFont(size=12))
            self.select_folder_btn.grid(row=3, column=0, padx=20, pady=5)

            self.help_btn = ctk.CTkButton(self.sidebar_frame, text="❓ Ajuda", command=self.show_help, font=ctk.CTkFont(size=12))
            self.help_btn.grid(row=4, column=0, padx=20, pady=5)

            # NOVO BOTÃO DE LIMPEZA
            self.clean_btn = ctk.CTkButton(self.sidebar_frame, text="🧹 Limpar Dashboard", command=self.start_cleaning, 
                                          font=ctk.CTkFont(size=12), fg_color="#FF9800", hover_color="#F57C00")
            self.clean_btn.grid(row=5, column=0, padx=20, pady=5)

            import platform
            self.system_label = ctk.CTkLabel(self.sidebar_frame, text=f"Sistema: {platform.system()}", font=ctk.CTkFont(size=10))
            self.system_label.grid(row=7, column=0, padx=20, pady=10)

            self.main_frame = ctk.CTkFrame(self, corner_radius=10)
            self.main_frame.grid(row=0, column=1, sticky="nsew", padx=20, pady=20)
            self.main_frame.grid_columnconfigure(0, weight=1)
            self.main_frame.grid_rowconfigure(1, weight=1)

            self.main_title = ctk.CTkLabel(self.main_frame, text="Sistema de Transporte de Dados", font=ctk.CTkFont(size=20, weight="bold"))
            self.main_title.grid(row=0, column=0, padx=20, pady=20)

            self.status_frame = ctk.CTkFrame(self.main_frame)
            self.status_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=(0, 20))
            self.status_frame.grid_columnconfigure(0, weight=1)
            self.status_frame.grid_rowconfigure(1, weight=1)

            self.info_text = ctk.CTkTextbox(self.status_frame, height=200)
            self.info_text.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
            self.info_text.configure(state="disabled")

            self.progress_frame = ctk.CTkFrame(self.status_frame)
            self.progress_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=10)
            self.progress_frame.grid_columnconfigure(0, weight=1)

            self.progress_label = ctk.CTkLabel(self.progress_frame, text="Pronto para processar", font=ctk.CTkFont(size=12))
            self.progress_label.grid(row=0, column=0, sticky="w", padx=10, pady=5)

            self.progress_bar = ctk.CTkProgressBar(self.progress_frame)
            self.progress_bar.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 10))
            self.progress_bar.set(0)

            self.stats_frame = ctk.CTkFrame(self.main_frame)
            self.stats_frame.grid(row=2, column=0, sticky="ew", padx=20, pady=(0, 20))
            self.stats_frame.grid_columnconfigure((0, 1, 2), weight=1)

            self.cities_label = ctk.CTkLabel(self.stats_frame, text="Cidades: 0", font=ctk.CTkFont(size=14, weight="bold"))
            self.cities_label.grid(row=0, column=0, padx=20, pady=15)

            self.updated_label = ctk.CTkLabel(self.stats_frame, text="Atualizadas: 0", font=ctk.CTkFont(size=14, weight="bold"))
            self.updated_label.grid(row=0, column=1, padx=20, pady=15)

            self.categories_label = ctk.CTkLabel(self.stats_frame, text="Categorias: 0", font=ctk.CTkFont(size=14, weight="bold"))
            self.categories_label.grid(row=0, column=2, padx=20, pady=15)

        def select_folder(self):
            from tkinter import filedialog, messagebox
            folder = filedialog.askdirectory(title="Selecione a pasta com os arquivos")
            if folder:
                self.base_dir = Path(folder)
                self.final_path = self.base_dir / self.final_name
                self.update_initial_info()
                messagebox.showinfo("Sucesso", f"Pasta selecionada:\n{self.base_dir}")

        def show_help(self):
            from tkinter import messagebox
            messagebox.showinfo("Ajuda - Athena Dashboard",
"""🎯 MANUAL RÁPIDO ATHENA DASHBOARD

1) Coloque na mesma pasta:
   - DASHBOARDFINAL.xlsx
   - JoãoPessoa.xls/.xlsx, SãoPaulo.xls/.xlsx, etc.

2) Clique em 'Processar Dados'

Dica: se alguma planilha der erro de corrupção, abra no Excel/LibreOffice e 'Salvar como' .xlsx, depois rode de novo.""")

        def update_initial_info(self):
            import platform, glob, os
            self.info_text.configure(state="normal"); self.info_text.delete("1.0", "end")
            info = [
                "📊 SISTEMA DE TRANSPORTE DE DADOS",
                "=" * 50,
                f"📁 Pasta: {self.base_dir}",
                f"💻 Sistema: {platform.system()}",
                "",
                "🔍 ARQUIVOS ENCONTRADOS:"
            ]
            if not self.final_path.exists():
                info.append("❌ DASHBOARDFINAL.xlsx NÃO ENCONTRADO!")
            else:
                info.append("✅ DASHBOARDFINAL.xlsx - OK")
            city_files = self.find_city_files()
            for f in city_files:
                info.append(f"   📄 {os.path.basename(f)}")
            if not city_files:
                info.append("   ⚠️  Nenhuma planilha de cidade encontrada")
            info += ["", "🎯 INSTRUÇÕES:", "1. Clique em 'Processar Dados'", "2. Aguarde", "3. Backup é automático",
                     "", "✅ PRONTO" if self.final_path.exists() else "❌ FALTA DASHBOARDFINAL.xlsx"]
            self.info_text.insert("1.0", "\n".join(info))
            self.info_text.configure(state="disabled")
            self.cities_label.configure(text=f"Cidades: {len(city_files)}")

        def find_city_files(self):
            import glob, os
            patterns = ["*.xls", "*.xlsx"]
            files = []
            for p in patterns:
                files.extend(glob.glob(str(self.base_dir / p)))
            return [f for f in files
                    if os.path.basename(f) != self.final_name
                    and not os.path.basename(f).startswith("~$")
                    and not os.path.basename(f).startswith("DASHBOARDFINAL_backup_")]

        def start_processing(self):
            from tkinter import messagebox
            import threading
            if self.is_processing:
                return
            if not self.final_path.exists():
                messagebox.showerror("Erro", f"Arquivo não encontrado:\n{self.final_path}\n\nColoque o arquivo DASHBOARDFINAL.xlsx na pasta.")
                return
            self.is_processing = True
            self.process_btn.configure(state="disabled")
            self.progress_bar.set(0)
            self.info_text.configure(state="normal")
            self.info_text.delete("1.0", "end")
            self.info_text.insert("1.0", "🚀 INICIANDO PROCESSAMENTO...\n\n")
            self.info_text.configure(state="disabled")
            threading.Thread(target=self.process_data, daemon=True).start()

        def finish_processing(self, success):
            """Finaliza o processamento e atualiza a UI"""
            def update_ui():
                self.is_processing = False
                self.process_btn.configure(state="normal")
                if success:
                    self.progress_bar.set(1.0)
                    self.progress_label.configure(text="Processamento concluído!")
                    # Mostrar mensagem de sucesso
                    self.log_message("\n🎉 PROCESSAMENTO FINALIZADO COM SUCESSO!")
                else:
                    self.progress_label.configure(text="Processamento falhou!")
                    self.log_message("\n💥 PROCESSAMENTO FALHOU!")
            
            # Agendar a atualização na thread principal
            self.after(0, update_ui)

        def process_data(self):
            import os, shutil
            from openpyxl import load_workbook
            from datetime import datetime

            try:
                self.log_message("📁 Criando backup do dashboard...")
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_path = self.base_dir / f"DASHBOARDFINAL_backup_{ts}.xlsx"
                if self.final_path.exists():
                    shutil.copyfile(self.final_path, backup_path)
                    self.log_message(f"✅ Backup criado: {backup_path.name}")
                else:
                    self.log_message("❌ Dashboard não encontrado!")
                    self.finish_processing(False); return

                self.log_message("📊 Carregando dashboard...")
                try:
                    wb = load_workbook(self.final_path)
                    self.log_message(f"✅ Dashboard carregado - {len(wb.sheetnames)} abas")
                except Exception as e:
                    self.log_message(f"❌ Erro ao carregar dashboard: {e}")
                    self.finish_processing(False); return

                city_files = self.find_city_files()
                self.log_message(f"🏙️  {len(city_files)} arquivos de cidades encontrados")
                if not city_files:
                    self.log_message("⚠️  Nenhuma planilha de cidade para processar")
                    self.finish_processing(True); return

                total_updated = 0; total_categories = 0
                for i, city_file in enumerate(city_files):
                    self.update_progress((i/len(city_files))*100, f"Processando {os.path.basename(city_file)}...")
                    city_name = os.path.splitext(os.path.basename(city_file))[0]
                    self.log_message(f"\n📋 PROCESSANDO: {city_name}")
                    try:
                        expenses = self.extract_expenses_vertical(city_file)
                        if expenses:
                            updated = self.update_dashboard_city_sheet(wb, city_name, expenses)
                            total_updated += updated
                            total_categories += len(expenses)
                            self.log_message(f"✅ {city_name}: {len(expenses)} categorias, {updated} atualizadas")
                        else:
                            self.log_message(f"⚠️  {city_name}: Nenhum dado extraído")
                    except RuntimeError as e:
                        # erro de leitura irrecuperável (ex.: corrupção)
                        self.log_message(f"❌ {city_name}: {e}")
                    except Exception as e:
                        self.log_message(f"❌ Erro em {city_name}: {e}")

                self.update_progress(90, "Salvando dashboard...")
                try:
                    wb.save(self.final_path)
                    
                    # Verificar se o arquivo foi realmente modificado
                    file_size = os.path.getsize(self.final_path)
                    self.log_message(f"💾 Dashboard salvo com sucesso! Tamanho: {file_size} bytes")
                    
                except Exception as e:
                    self.log_message(f"❌ Erro ao salvar: {e}")
                    # Tentativa alternativa de salvamento
                    try:
                        backup_name = f"DASHBOARDFINAL_emergency_backup.xlsx"
                        wb.save(self.base_dir / backup_name)
                        self.log_message(f"⚠️  Backup de emergência salvo como: {backup_name}")
                    except Exception as e2:
                        self.log_message(f"❌ Falha no backup de emergência: {e2}")
                    self.finish_processing(False)
                    return

                self.after(0, lambda: self.update_final_stats(total_updated, total_categories, len(city_files)))
                self.update_progress(100, "Processamento concluído!")
                self.log_message("\n🎯 PROCESSAMENTO CONCLUÍDO!")
                self.log_message(f"📊 Resumo: {len(city_files)} cidades, {total_categories} categorias, {total_updated} atualizações")
                self.finish_processing(True)

            except Exception as e:
                self.log_message(f"❌ ERRO CRÍTICO: {e}")
                self.finish_processing(False)

        def log_message(self, message):
            def update_text():
                self.info_text.configure(state="normal")
                self.info_text.insert("end", message + "\n")
                self.info_text.see("end")
                self.info_text.configure(state="disabled")
            self.after(0, update_text)

        def update_progress(self, value, text):
            def update():
                self.progress_bar.set(value / 100)
                self.progress_label.configure(text=text)
            self.after(0, update)

        def update_final_stats(self, updated, categories, cities):
            self.updated_label.configure(text=f"Atualizadas: {updated}")
            self.categories_label.configure(text=f"Categorias: {categories}")
            self.cities_label.configure(text=f"Cidades: {cities}")

        def similarity_score(self, str1, str2):
            """Calcula similaridade entre strings (0.0 a 1.0)"""
            return SequenceMatcher(None, str1, str2).ratio()

        # --------- EXTRAÇÃO (vertical com normalização) ---------
        def extract_expenses_vertical(self, file_path):
            import pandas as pd
            expenses = {}
            df = read_excel_any(Path(file_path))
            if df.shape[1] < 2:
                self.log_message("  ⚠️ Planilha sem ao menos 2 colunas (descrição/valor)")
                return expenses

            # Log da estrutura da planilha
            self.log_message(f"  📊 Estrutura da planilha: {df.shape[0]} linhas x {df.shape[1]} colunas")
            
            for idx, row in df.iterrows():
                desc_raw = row.iloc[0] if df.shape[1] >= 1 else None
                val_raw  = row.iloc[1] if df.shape[1] >= 2 else None
                
                if pd.isna(desc_raw) or pd.isna(val_raw):
                    continue
                    
                valor = to_float(val_raw)
                if valor is None:
                    continue

                desc_norm = norm_text(desc_raw)
                matched = False
                
                # Busca DIRETA no mapeamento - sem verificações complexas
                for cat_src_norm, cat_dst in CATEGORY_MAPPING.items():
                    # Verifica se a descrição normalizada contém a categoria normalizada
                    if cat_src_norm in desc_norm:
                        # Verifica se esta categoria já foi processada (evita duplicação)
                        if cat_dst not in expenses:
                            expenses[cat_dst] = valor
                            self.log_message(f"  ✅ {cat_dst}: R$ {valor:,.2f} (origem: '{desc_raw}')")
                            matched = True
                            break
                        else:
                            self.log_message(f"  ⚠️ Categoria duplicada ignorada: {cat_dst}")
                            matched = True
                            break
                
                # Se não encontrou no mapping, log para debug posterior
                if not matched:
                    self.log_message(f"  🔍 Categoria não mapeada: '{desc_raw}' → R$ {valor:,.2f}")
                        
            return expenses

        # --------- MATCH DA ABA (normalizado) ---------
        def find_city_sheet(self, wb, city_name):
            wanted = norm_text(city_name)
            for sheet_name in wb.sheetnames:
                if norm_text(sheet_name).find(wanted) != -1:
                    return sheet_name
            return None

        def update_dashboard_city_sheet(self, wb, city_name, expenses):
            sheet_name = self.find_city_sheet(wb, city_name)
            if not sheet_name:
                self.log_message(f"❌ Aba para '{city_name}' não encontrada")
                return 0
            
            ws = wb[sheet_name]
            updated_count = 0
            matched_categories = set()
            used_expenses = set()  # Controla quais despesas já foram utilizadas
            
            self.log_message(f"  🔍 Procurando {len(expenses)} categorias na aba '{sheet_name}'")
            
            for row in range(1, ws.max_row + 1):
                cell_a = ws[f'A{row}']
                if cell_a.value and isinstance(cell_a.value, str):
                    categoria_dashboard = cell_a.value.strip()
                    dash_norm = norm_text(categoria_dashboard)
                    
                    if categoria_dashboard in matched_categories:
                        continue
                        
                    best_match = None
                    best_similarity = 0
                    
                    for cat_extraida, valor in expenses.items():
                        # Pula se esta despesa já foi usada antes
                        if cat_extraida in used_expenses:
                            continue
                            
                        cat_extraida_norm = norm_text(cat_extraida)
                        similarity = self.similarity_score(cat_extraida_norm, dash_norm)
                        
                        # CORREÇÃO: Matching mais específico para tarifas bancárias
                        if 'tarifas bancarias' in dash_norm or 'tarifas bancarias' in cat_extraida_norm:
                            # Para tarifas, verificar matching exato do tipo (se perfeito)
                            if 'ted' in dash_norm and 'ted' in cat_extraida_norm:
                                similarity = 1.0  
                            elif 'pix' in dash_norm and 'pix' in cat_extraida_norm:
                                similarity = 1.0  
                            elif 'boletos' in dash_norm and 'boletos' in cat_extraida_norm:
                                similarity = 1.0  
                            elif 'cartao' in dash_norm and 'cartao' in cat_extraida_norm:
                                similarity = 1.0 
                            else:
                                # Se não for matching específico, vai reduzir similaridade
                                similarity = similarity * 0.5
                        
                        # Similaridade mínima ajustada
                        min_similarity = 0.8
                        
                        if similarity > best_similarity and similarity > min_similarity:
                            best_similarity = similarity
                            best_match = (cat_extraida, valor)
                    
                    if best_match and best_similarity > 0.8:
                        cat_extraida, valor = best_match
                        cell_b = ws[f'B{row}']
                        old_value = cell_b.value
                        
                        if old_value != valor:
                            cell_b.value = valor
                            updated_count += 1
                            matched_categories.add(categoria_dashboard)
                            used_expenses.add(cat_extraida)  # Marca como usada
                            
                            if cell_b.value == valor:
                                self.log_message(f"  🔄 {categoria_dashboard}: R$ {valor:,.2f} (era: {old_value}) - Similaridade: {best_similarity:.2f}")
                            else:
                                self.log_message(f"  ❌ Falha ao salvar {categoria_dashboard}")
                        else:
                            self.log_message(f"  ✅ {categoria_dashboard}: R$ {valor:,.2f} (já atualizado) - Similaridade: {best_similarity:.2f}")
            
            # Log das categorias que não foram encontradas
            unmatched = set(expenses.keys()) - used_expenses
            if unmatched:
                self.log_message(f"  ⚠️ Categorias não encontradas no dashboard: {unmatched}")
            
            self.log_message(f"  📝 Total de células atualizadas: {updated_count}")
            return updated_count

        # ========== FUNÇÕES DE LIMPEZA ==========
        def start_cleaning(self):
            """Inicia o processo de limpeza do dashboard"""
            from tkinter import messagebox
            import threading
            
            if self.is_processing:
                messagebox.showwarning("Atenção", "Já existe um processamento em andamento!")
                return
                
            if not self.final_path.exists():
                messagebox.showerror("Erro", f"Arquivo não encontrado:\n{self.final_path}")
                return
                
            confirm = messagebox.askyesno("Confirmar Limpeza", 
                                         "⚠️ ATENÇÃO: Esta ação irá limpar TODOS os valores das células de despesas em TODAS as abas (exceto a primeira) do dashboard.\n\n"
                                         "Deseja continuar?")
            if not confirm:
                return
                
            self.is_processing = True
            self.clean_btn.configure(state="disabled")
            self.process_btn.configure(state="disabled")
            self.progress_bar.set(0)
            
            self.info_text.configure(state="normal")
            self.info_text.delete("1.0", "end")
            self.info_text.insert("1.0", "🧹 INICIANDO LIMPEZA DO DASHBOARD...\n\n")
            self.info_text.configure(state="disabled")
            
            threading.Thread(target=self.clean_dashboard, daemon=True).start()

        def clean_dashboard(self):
            """Limpa os ranges específicos nas abas a partir da segunda"""
            import os, shutil
            from openpyxl import load_workbook
            from datetime import datetime

            try:
                self.log_message("📁 Criando backup antes da limpeza...")
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_path = self.base_dir / f"DASHBOARDFINAL_backup_antes_da_limpeza_{ts}.xlsx"
                if self.final_path.exists():
                    shutil.copyfile(self.final_path, backup_path)
                    self.log_message(f"✅ Backup criado: {backup_path.name}")
                else:
                    self.log_message("❌ Dashboard não encontrado!")
                    self.finish_cleaning(False)
                    return

                self.log_message("📊 Carregando dashboard para limpeza...")
                try:
                    wb = load_workbook(self.final_path)
                    self.log_message(f"✅ Dashboard carregado - {len(wb.sheetnames)} abas")
                except Exception as e:
                    self.log_message(f"❌ Erro ao carregar dashboard: {e}")
                    self.finish_cleaning(False)
                    return

                # Ranges a serem limpos (conforme especificado)
                ranges_to_clear = [
                    "B4:B15",    # Pessoal
                    "B18:B31",   # Administrativas  
                    "B34:B38",   # Tributos e Taxas
                    "B41:B43",   # Serviços Profissionais e Legais
                    "B46",       # Movimentações com Sócios
                    "B49:B54",   # Financeiras
                    "B57",       # Marketing e Comunicação
                    "B59:B63",   # Investimentos e Manutenções
                    "B67"        # Propaganda e Publicidade
                ]
                
                total_cleaned = 0
                # Pega apenas as abas a partir da segunda (índice 1 em diante)
                sheets_to_clean = wb.sheetnames[1:]
                total_sheets = len(sheets_to_clean)
                
                if total_sheets == 0:
                    self.log_message("⚠️ Nenhuma aba para limpar (apenas a primeira aba encontrada)")
                    self.finish_cleaning(True)
                    return
                
                self.log_message(f"🔧 Limpando {total_sheets} abas (a partir da segunda)")
                
                for i, sheet_name in enumerate(sheets_to_clean):
                    self.update_progress((i/total_sheets)*100, f"Limpando aba: {sheet_name}...")
                    self.log_message(f"\n📋 LIMPANDO: {sheet_name}")
                    
                    ws = wb[sheet_name]
                    sheet_cleaned = 0
                    
                    for range_str in ranges_to_clear:
                        try:
                            # Limpa o range específico
                            cleaned_in_range = self.clear_range(ws, range_str)
                            sheet_cleaned += cleaned_in_range
                            
                        except Exception as e:
                            self.log_message(f"  ⚠️ Erro ao limpar range {range_str}: {e}")
                    
                    total_cleaned += sheet_cleaned
                    self.log_message(f"  ✅ {sheet_name}: {sheet_cleaned} células limpas")

                self.update_progress(90, "Salvando dashboard limpo...")
                try:
                    wb.save(self.final_path)
                    self.log_message(f"💾 Dashboard salvo com sucesso!")
                    
                except Exception as e:
                    self.log_message(f"❌ Erro ao salvar: {e}")
                    self.finish_cleaning(False)
                    return

                self.update_progress(100, "Limpeza concluída!")
                self.log_message("\n🎯 LIMPEZA CONCLUÍDA!")
                self.log_message(f"📊 Resumo: {total_sheets} abas processadas, {total_cleaned} células limpas no total")
                self.finish_cleaning(True)

            except Exception as e:
                self.log_message(f"❌ ERRO CRÍTICO: {e}")
                self.finish_cleaning(False)

        def clear_range(self, worksheet, range_str):
            """Limpa os valores em um range específico da planilha"""
            cleaned_count = 0
            
            # Converte o range string em coordenadas
            if ':' in range_str:
                # Range como "B4:B15"
                start_cell, end_cell = range_str.split(':')
                cells = worksheet[start_cell:end_cell]
                
                for row in cells:
                    for cell in row:
                        if cell.value is not None:
                            cell.value = None
                            cleaned_count += 1
            else:
                # Célula única como "B46"
                cell = worksheet[range_str]
                if cell.value is not None:
                    cell.value = None
                    cleaned_count += 1
                    
            return cleaned_count

        def finish_cleaning(self, success):
            """Finaliza o processo de limpeza"""
            def update_ui():
                self.is_processing = False
                self.clean_btn.configure(state="normal")
                self.process_btn.configure(state="normal")
                if success:
                    self.progress_bar.set(1.0)
                    self.progress_label.configure(text="Limpeza concluída!")
                    self.log_message("\n🎉 DASHBOARD LIMPO COM SUCESSO!")
                else:
                    self.progress_label.configure(text="Limpeza falhou!")
                    self.log_message("\n💥 FALHA NA LIMPEZA!")
            
            self.after(0, update_ui)

    import customtkinter as ctk
    app = DashboardApp()
    app.mainloop()

if __name__ == "__main__":
    main()